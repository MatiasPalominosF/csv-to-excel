import os
import time
import threading
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from tqdm import tqdm

def adjust_column_width(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def refresh_bar(pbar):
    while not pbar.disable and pbar.n < pbar.total:
        time.sleep(1)
        pbar.refresh()

def convert_csv_to_excel(folder_path, output_folder):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    archivos_csv = [f for f in os.listdir(folder_path) if f.endswith(".csv")]

    with tqdm(
        total=len(archivos_csv),
        desc="📊 Convirtiendo CSVs",
        unit="archivo",
        bar_format="{l_bar}{bar}| {percentage:3.0f}% • {n_fmt}/{total_fmt} archivos • ⏱ {elapsed}"
    ) as pbar:
        # Iniciar hilo para refrescar la barra cada segundo
        refrescador = threading.Thread(target=refresh_bar, args=(pbar,))
        refrescador.start()

        for filename in archivos_csv:
            csv_file = os.path.join(folder_path, filename)
            excel_file = os.path.join(output_folder, os.path.splitext(filename)[0] + '.xlsx')

            df = pd.read_csv(csv_file, low_memory=False)
            pbar.set_postfix(archivo=filename[:30], filas=len(df))
            df.to_excel(excel_file, index=False)

            wb = load_workbook(excel_file)
            ws = wb.active
            adjust_column_width(ws)

            for row in ws.iter_rows():
                ws.row_dimensions[row[0].row].auto_size = True

            wb.save(excel_file)
            pbar.update(1)

        refrescador.join()  # Esperar que termine el hilo

    print("🎉 ¡Archivos convertidos correctamente!")

# Directorio que contiene los archivos CSV
csv_directory = 'archivos/'

# Carpeta donde se guardarán los archivos Excel convertidos
output_folder = 'archivos_excel/'

# Llama a la función para convertir los archivos
convert_csv_to_excel(csv_directory, output_folder)
